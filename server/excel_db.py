"""Excel-backed data layer (openpyxl). This is explicitly a TEMPORARY interim
layer per the project brief -- single workbook, single sheet, append-only, no
schema versioning or write-locking. Only qualifying scrap records are written
by default (config.LOG_DISCARDS can turn on optional discard logging for POC
debugging without changing that default spec-literal behavior).
"""

import json
import os
from datetime import datetime

from openpyxl import Workbook, load_workbook

import config

HEADERS = [
    "餘料編號", "原始設計圖來源/批次", "原始檔名", "來源類型", "輪廓判定方式",
    "總面積(cm²)", "形狀描述/輪廓座標", "可容納A的數量", "可容納B的數量",
    "其他標準品容納數量(JSON)", "材質", "使用格點大小(cm)", "判定結果", "建檔時間",
]

MAX_POINTS_STORED = 100


def ensure_workbook(path=None):
    path = path or config.EXCEL_DB_PATH
    os.makedirs(os.path.dirname(path), exist_ok=True)
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "餘料資料庫"
        ws.append(HEADERS)
        wb.save(path)
    return path


def append_record(
    *, batch, source_filename, source_type, contour_method, area_cm2,
    polygon_cm, length_cm, width_cm, fits_by_product, material,
    grid_cell_size_cm, verdict, path=None,
    sheet_name=None, sheet_area_cm2=None, product_area_cm2=None, product_fit_count=None,
):
    path = ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb.active

    existing_data_rows = ws.max_row - 1  # row 1 is the header
    scrap_id = f"SCRAP-{existing_data_rows + 1:05d}"

    fits_a = (fits_by_product.get("A") or {}).get("fits", 0)
    fits_b = (fits_by_product.get("B") or {}).get("fits", 0)
    other_fits = {
        pid: r.get("fits", 0)
        for pid, r in fits_by_product.items()
        if pid not in ("A", "B")
    }
    shape_desc = {
        "length_cm": length_cm,
        "width_cm": width_cm,
        "points": (polygon_cm or [])[:MAX_POINTS_STORED],
    }
    # Design-to-leftover provenance (see qualification.evaluate_design): what sheet
    # and product area this leftover was computed from, not directly measured from.
    if sheet_name is not None or sheet_area_cm2 is not None or product_area_cm2 is not None:
        shape_desc["sheet_name"] = sheet_name
        shape_desc["sheet_area_cm2"] = sheet_area_cm2
        shape_desc["product_area_cm2"] = product_area_cm2
        shape_desc["product_fit_count"] = product_fit_count

    row = [
        scrap_id,
        batch or "",
        source_filename or "",
        source_type or "",
        contour_method or "",
        area_cm2,
        json.dumps(shape_desc, ensure_ascii=False),
        fits_a,
        fits_b,
        json.dumps(other_fits, ensure_ascii=False),
        material or "",
        grid_cell_size_cm,
        verdict,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ]
    ws.append(row)
    wb.save(path)

    return dict(zip(HEADERS, row))


def get_record(scrap_id, path=None):
    """Looks up one previously stocked-in row by 餘料編號. Returns a dict keyed
    by HEADERS, or None if not found (including when the workbook doesn't exist
    yet). Used by the SAP export endpoint to pull a scrap record's dimensions."""
    path = path or config.EXCEL_DB_PATH
    if not os.path.exists(path):
        return None
    wb = load_workbook(path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] == scrap_id:
            return dict(zip(HEADERS, row))
    return None
