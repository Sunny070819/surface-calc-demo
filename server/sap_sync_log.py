"""Local audit trail for SAP ICF sync attempts (see sap_client.py), so the
operator can reconcile against an SE16 lookup on ZAI_DIM_LOG without needing SAP
access for every check. Records the board/product-area values this software
sent, and whatever remnant_area/batch/bin_zone SAP hands back -- remnant_area
here is SAP's own "area" response field (see sap_client.post_remnant), which
empirically just echoes back the precomputed_area we sent rather than being
recomputed server-side. Same single-workbook/single-sheet POC pattern as
excel_db.py.
"""

import os
from datetime import datetime

from openpyxl import Workbook, load_workbook

import config

HEADERS = [
    "同步時間", "本地餘料編號", "source_id", "matnr",
    "傳送的board_length(cm)", "傳送的board_width(cm)", "傳送的product_area(cm²)",
    "候選方案(JSON)", "綁定狀態", "HTTP狀態碼", "SAP狀態(S/E)", "SAP回傳guid", "SAP回傳batch", "SAP回傳bin_zone",
    "SAP回傳remnant_area", "SAP回傳message", "結果", "錯誤訊息",
]


def ensure_workbook(path=None):
    path = path or config.SAP_SYNC_LOG_PATH
    os.makedirs(os.path.dirname(path), exist_ok=True)
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "SAP同步紀錄"
        ws.append(HEADERS)
        wb.save(path)
    return path


def append_sync_record(*, scrap_id, source_id, matnr, board_length_cm, board_width_cm,
                        product_area_cm2, result, path=None):
    """`result` is a dict as returned by sap_client.post_remnant."""
    path = ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb.active

    payload = result.get("payload") or {}
    row = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        scrap_id, source_id, matnr,
        board_length_cm, board_width_cm, product_area_cm2,
        payload.get("candidate_options") or "",
        payload.get("binding_status") or "UNBOUND",
        result.get("http_status"), result.get("sap_status"),
        result.get("guid"), result.get("batch"), result.get("bin_zone"),
        result.get("remnant_area"), result.get("message"),
        "成功" if result.get("ok") else "失敗",
        result.get("error") or "",
    ]
    ws.append(row)
    wb.save(path)
    return dict(zip(HEADERS, row))


# Separate sheet ("簡化匯出紀錄") in the same workbook for the simplified
# export path (sap_client.post_remnant_simple / app.py's
# /api/sap/export-simple/<scrap_id>) -- its payload has no board/product-area
# fields, so it doesn't fit HEADERS above; kept as its own sheet/schema rather
# than shoehorned into the existing one, and append_sync_record above is
# untouched by this addition.
SIMPLE_HEADERS = [
    "同步時間", "本地餘料編號", "matnr", "sloc", "quantity",
    "HTTP狀態碼", "SAP狀態(S/E)", "SAP回傳matdoc", "SAP回傳matdoc_year",
    "SAP回傳message", "結果", "錯誤訊息",
]


def ensure_simple_workbook(path=None):
    path = ensure_workbook(path)
    wb = load_workbook(path)
    if "簡化匯出紀錄" not in wb.sheetnames:
        ws = wb.create_sheet("簡化匯出紀錄")
        ws.append(SIMPLE_HEADERS)
        wb.save(path)
    return path


def append_simple_sync_record(*, scrap_id, matnr, sloc, quantity, result, path=None):
    """`result` is a dict as returned by sap_client.post_remnant_simple."""
    path = ensure_simple_workbook(path)
    wb = load_workbook(path)
    ws = wb["簡化匯出紀錄"]

    row = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        scrap_id, matnr, sloc, quantity,
        result.get("http_status"), result.get("sap_status"),
        result.get("matdoc"), result.get("matdoc_year"), result.get("message"),
        "成功" if result.get("ok") else "失敗",
        result.get("error") or "",
    ]
    ws.append(row)
    wb.save(path)
    return dict(zip(SIMPLE_HEADERS, row))
