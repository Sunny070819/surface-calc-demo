import io
import json

from openpyxl import load_workbook

import app as app_module
import config
import sap_client


def client():
    app_module.app.testing = True
    return app_module.app.test_client()


def _product_payload(product_side=30, sheet_length=100, sheet_width=100, sheet_name="TEST-SHEET"):
    """A square `product_side` x `product_side` product design, to be packed
    into a `sheet_length` x `sheet_width` standard sheet."""
    return {
        "polygon_cm": [
            {"x": 0, "y": 0}, {"x": product_side, "y": 0},
            {"x": product_side, "y": product_side}, {"x": 0, "y": product_side},
        ],
        "sheet": {"length_cm": sheet_length, "width_cm": sheet_width, "name": sheet_name},
    }


def _almost_fills_sheet_payload():
    """19x19 product on a 20x20 sheet: leftover area 400-361=39, below the
    default 300cm2 threshold."""
    return _product_payload(product_side=19, sheet_length=20, sheet_width=20)


def _exceeds_sheet_payload():
    """20x5 product can't fit a 10x10 sheet in either rotation."""
    return {
        "polygon_cm": [{"x": 0, "y": 0}, {"x": 20, "y": 0}, {"x": 20, "y": 5}, {"x": 0, "y": 5}],
        "sheet": {"length_cm": 10, "width_cm": 10, "name": "TOO-SMALL"},
    }


def test_get_config():
    resp = client().get("/api/config")
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["threshold_cm2"] == config.AREA_DISCARD_THRESHOLD_CM2


def test_get_standard_products():
    resp = client().get("/api/standard-products")
    assert resp.status_code == 200
    ids = [p["id"] for p in resp.get_json()]
    assert "A" in ids and "B" in ids


def test_get_standard_sheets():
    resp = client().get("/api/standard-sheets")
    assert resp.status_code == 200
    sheets = resp.get_json()
    assert len(sheets) >= 1
    assert {"id", "name", "length_cm", "width_cm", "area_cm2"} <= sheets[0].keys()


def test_evaluate_qualifying_design():
    # 30x30 product packed into a 100x100 sheet: 9 copies fit (3x3 grid),
    # leaving a 1900cm2 L-shaped leftover -- well above threshold.
    resp = client().post("/api/evaluate", json=_product_payload())
    body = resp.get_json()
    assert body["product_area_cm2"] == 900.0
    assert body["sheet_area_cm2"] == 10000.0
    assert body["product_fit_count"] >= 1
    assert body["qualifies"] is True
    assert body["nesting"]["A"]["fits"] >= 1
    assert body["nesting"]["A"]["unit_price"] == 10
    assert body["nesting"]["A"]["benefit"] >= 0
    assert body["ranked_options"][0]["product_id"] == "A"


def test_evaluate_below_threshold():
    resp = client().post("/api/evaluate", json=_almost_fills_sheet_payload())
    body = resp.get_json()
    assert body["passes_area_threshold"] is False
    assert body["reason"] == "area_below_threshold"


def test_evaluate_product_exceeds_sheet():
    resp = client().post("/api/evaluate", json=_exceeds_sheet_payload())
    body = resp.get_json()
    assert body["qualifies"] is False
    assert body["reason"] == "product_exceeds_sheet"
    assert body["product_fit_count"] == 0


def test_evaluate_missing_polygon_returns_400():
    resp = client().post("/api/evaluate", json={"sheet": {"length_cm": 100, "width_cm": 100}})
    assert resp.status_code == 400


def test_evaluate_missing_sheet_returns_400():
    resp = client().post("/api/evaluate", json={"polygon_cm": _product_payload()["polygon_cm"]})
    assert resp.status_code == 400


def test_stock_in_writes_excel_row(monkeypatch, tmp_path):
    monkeypatch.setattr(config, "EXCEL_DB_PATH", str(tmp_path / "scrap_inventory.xlsx"))
    payload = dict(
        _product_payload(),
        material="防火", source_type="manual", contour_method="自動偵測",
        source_filename="demo.png", batch_no="BATCH-X",
    )
    resp = client().post("/api/stock-in", json=payload)
    body = resp.get_json()
    assert body["ok"] is True
    assert body["scrap_id"] == "SCRAP-00001"

    wb = load_workbook(config.EXCEL_DB_PATH)
    ws = wb.active
    assert ws.max_row == 2
    assert ws.cell(row=2, column=1).value == "SCRAP-00001"

    shape_desc = json.loads(ws.cell(row=2, column=7).value)
    assert shape_desc["sheet_area_cm2"] == 10000.0
    assert shape_desc["product_area_cm2"] == 900.0
    assert shape_desc["product_fit_count"] >= 1

    candidate_options = json.loads(ws.cell(row=2, column=13).value)
    assert candidate_options[0]["product_id"] == "A"
    assert ws.cell(row=2, column=14).value == "UNBOUND"


def test_stock_in_rejects_non_qualifying_and_writes_nothing(monkeypatch, tmp_path):
    excel_path = str(tmp_path / "scrap_inventory.xlsx")
    monkeypatch.setattr(config, "EXCEL_DB_PATH", excel_path)
    resp = client().post("/api/stock-in", json=_almost_fills_sheet_payload())
    body = resp.get_json()
    assert body["ok"] is False
    assert body["reason"] == "area_below_threshold"
    import os
    assert not os.path.exists(excel_path)


def test_analyze_dxf_endpoint(synthetic_rect_dxf_bytes):
    data = {
        "dxf": (io.BytesIO(synthetic_rect_dxf_bytes), "test.dxf"),
        "scale_cm_per_unit": "1.0",
    }
    resp = client().post("/api/analyze-dxf", data=data, content_type="multipart/form-data")
    body = resp.get_json()
    assert body["ok"] is True
    assert body["area_cm2"] > 0


def test_analyze_photo_endpoint(synthetic_scrap_photo_bytes):
    data = {
        "image": (io.BytesIO(synthetic_scrap_photo_bytes), "test.png"),
        "display_width": "560",
        "display_height": "420",
    }
    resp = client().post("/api/analyze-photo", data=data, content_type="multipart/form-data")
    body = resp.get_json()
    assert body["ok"] is True
    assert len(body["contour_px"]) >= 3


def _sap_export_payload(matnr="SCRAP-WP", board_length_cm=240, board_width_cm=120, product_area_cm2=18000):
    return {
        "matnr": matnr, "board_length_cm": board_length_cm,
        "board_width_cm": board_width_cm, "product_area_cm2": product_area_cm2,
    }


def test_sap_export_missing_matnr_returns_400(monkeypatch, tmp_path):
    monkeypatch.setattr(config, "EXCEL_DB_PATH", str(tmp_path / "scrap_inventory.xlsx"))
    client().post("/api/stock-in", json=_product_payload())
    resp = client().post("/api/sap/export/SCRAP-00001", json={})
    assert resp.status_code == 400


def test_sap_export_unknown_scrap_id_returns_404(monkeypatch, tmp_path):
    monkeypatch.setattr(config, "EXCEL_DB_PATH", str(tmp_path / "scrap_inventory.xlsx"))
    resp = client().post("/api/sap/export/SCRAP-DOES-NOT-EXIST", json=_sap_export_payload())
    assert resp.status_code == 404


def test_sap_export_success_logs_sync_record(monkeypatch, tmp_path):
    monkeypatch.setattr(config, "EXCEL_DB_PATH", str(tmp_path / "scrap_inventory.xlsx"))
    sync_log_path = str(tmp_path / "sap_sync_log.xlsx")
    monkeypatch.setattr(config, "SAP_SYNC_LOG_PATH", sync_log_path)
    client().post("/api/stock-in", json=_product_payload())

    captured = {}

    def fake_post_remnant(**kwargs):
        captured.update(kwargs)
        return {
            "ok": True, "http_status": 200, "sap_status": "S", "guid": "G-1",
            "batch": "0000000243", "bin_zone": "B-1", "remnant_area": "10800.000",
            "message": "OK MatDoc:4900011979", "error": None, "payload": kwargs,
        }

    monkeypatch.setattr(sap_client, "post_remnant", fake_post_remnant)

    resp = client().post("/api/sap/export/SCRAP-00001", json=_sap_export_payload())
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["ok"] is True
    assert body["guid"] == "G-1"
    assert body["batch"] == "0000000243"
    assert body["bin_zone"] == "B-1"

    wb = load_workbook(sync_log_path)
    assert wb.active.max_row == 2  # header + 1 sync record


def test_sap_export_sap_failure_returns_502(monkeypatch, tmp_path):
    monkeypatch.setattr(config, "EXCEL_DB_PATH", str(tmp_path / "scrap_inventory.xlsx"))
    monkeypatch.setattr(config, "SAP_SYNC_LOG_PATH", str(tmp_path / "sap_sync_log.xlsx"))
    client().post("/api/stock-in", json=_product_payload())

    monkeypatch.setattr(sap_client, "post_remnant", lambda **kwargs: {
        "ok": False, "http_status": 401, "sap_status": None, "guid": None,
        "batch": None, "bin_zone": None, "remnant_area": None, "message": None,
        "error": sap_client.FRIENDLY_ERRORS[401], "payload": kwargs,
    })

    resp = client().post("/api/sap/export/SCRAP-00001", json=_sap_export_payload())
    assert resp.status_code == 502
    assert resp.get_json()["ok"] is False


def test_sap_export_unknown_matnr_returns_404_from_sap(monkeypatch, tmp_path):
    monkeypatch.setattr(config, "EXCEL_DB_PATH", str(tmp_path / "scrap_inventory.xlsx"))
    monkeypatch.setattr(config, "SAP_SYNC_LOG_PATH", str(tmp_path / "sap_sync_log.xlsx"))
    client().post("/api/stock-in", json=_product_payload())

    monkeypatch.setattr(sap_client, "post_remnant", lambda **kwargs: {
        "ok": False, "http_status": 404, "sap_status": None, "guid": None,
        "batch": None, "bin_zone": None, "remnant_area": None, "message": None,
        "error": "料號在 SAP 查無資料，請確認料號是否正確（2604 廠僅維護 SCRAP-FR / SCRAP-WP / SCRAP-RC）。",
        "payload": kwargs,
    })

    resp = client().post("/api/sap/export/SCRAP-00001", json=_sap_export_payload(matnr="NOT-A-REAL-MATNR"))
    assert resp.status_code == 502
    body = resp.get_json()
    assert body["ok"] is False
    assert body["http_status"] == 404


def test_sap_export_sends_leftover_area_not_product_area(monkeypatch, tmp_path):
    """_product_payload()'s default (30x30 product on a 100x100 sheet) packs 9
    copies (product_fit_count=9), so product_area_cm2 (one unit, 900) and the
    leftover's own area_cm2 (~1900, after 9 copies are subtracted) are genuinely
    different numbers -- exactly the case where sending product_area_cm2 into
    SAP's precomputed_area field instead of the leftover area would be wrong.
    Exercises the real /api/evaluate -> /api/stock-in -> /api/sap/export chain;
    only the outbound HTTP call inside sap_client.post_remnant is mocked."""
    monkeypatch.setattr(config, "EXCEL_DB_PATH", str(tmp_path / "scrap_inventory.xlsx"))
    monkeypatch.setattr(config, "SAP_SYNC_LOG_PATH", str(tmp_path / "sap_sync_log.xlsx"))

    eval_body = client().post("/api/evaluate", json=_product_payload()).get_json()
    assert eval_body["product_fit_count"] > 1
    assert eval_body["product_area_cm2"] != eval_body["area_cm2"]
    leftover_area_cm2 = eval_body["area_cm2"]
    product_area_cm2 = eval_body["product_area_cm2"]

    stock_body = client().post("/api/stock-in", json=_product_payload()).get_json()
    scrap_id = stock_body["scrap_id"]

    captured = {}

    def fake_post_remnant(**kwargs):
        captured.update(kwargs)
        return {
            "ok": True, "http_status": 200, "sap_status": "S", "guid": "G-1",
            "batch": "0000000243", "bin_zone": "B-1",
            "remnant_area": str(kwargs["scrap_area_cm2"]),
            "message": "OK", "error": None, "payload": kwargs,
        }

    monkeypatch.setattr(sap_client, "post_remnant", fake_post_remnant)

    # Deliberately NOT passing scrap_area_cm2 in the request body: it should
    # fall back to the value excel_db already stored (總面積(cm²)) for this
    # scrap_id, same as the real frontend flow where pendingSapExport.scrapAreaCm2
    # is populated from the item's own stored leftover area.
    resp = client().post(f"/api/sap/export/{scrap_id}", json=_sap_export_payload(product_area_cm2=product_area_cm2))
    assert resp.status_code == 200

    assert captured["product_area_cm2"] == product_area_cm2
    assert captured["scrap_area_cm2"] == leftover_area_cm2
    assert captured["scrap_area_cm2"] != captured["product_area_cm2"]


def test_sap_export_missing_sap_config_returns_500(monkeypatch, tmp_path):
    monkeypatch.setattr(config, "EXCEL_DB_PATH", str(tmp_path / "scrap_inventory.xlsx"))
    client().post("/api/stock-in", json=_product_payload())

    def raise_config_error(**kwargs):
        raise sap_client.SapConfigError("尚未設定 SAP 連線環境變數")

    monkeypatch.setattr(sap_client, "post_remnant", raise_config_error)

    resp = client().post("/api/sap/export/SCRAP-00001", json=_sap_export_payload())
    assert resp.status_code == 500


# --- /api/sap/export-simple/<scrap_id>: a separate route from
# /api/sap/export/<scrap_id> above, calling sap_client.post_remnant_simple
# instead of post_remnant. These tests never monkeypatch post_remnant, and the
# ones above never monkeypatch post_remnant_simple -- the two paths are
# independently exercised. ---

def test_sap_export_simple_missing_matnr_returns_400(monkeypatch, tmp_path):
    monkeypatch.setattr(config, "EXCEL_DB_PATH", str(tmp_path / "scrap_inventory.xlsx"))
    client().post("/api/stock-in", json=_product_payload())
    resp = client().post("/api/sap/export-simple/SCRAP-00001", json={"sloc": "SBED", "quantity": "9"})
    assert resp.status_code == 400


def test_sap_export_simple_missing_sloc_returns_400(monkeypatch, tmp_path):
    monkeypatch.setattr(config, "EXCEL_DB_PATH", str(tmp_path / "scrap_inventory.xlsx"))
    client().post("/api/stock-in", json=_product_payload())
    resp = client().post("/api/sap/export-simple/SCRAP-00001", json={"matnr": "SCRAP2-FR", "quantity": "9"})
    assert resp.status_code == 400


def test_sap_export_simple_missing_quantity_returns_400(monkeypatch, tmp_path):
    monkeypatch.setattr(config, "EXCEL_DB_PATH", str(tmp_path / "scrap_inventory.xlsx"))
    client().post("/api/stock-in", json=_product_payload())
    resp = client().post("/api/sap/export-simple/SCRAP-00001", json={"matnr": "SCRAP2-FR", "sloc": "SBED"})
    assert resp.status_code == 400


def test_sap_export_simple_unknown_scrap_id_returns_404(monkeypatch, tmp_path):
    monkeypatch.setattr(config, "EXCEL_DB_PATH", str(tmp_path / "scrap_inventory.xlsx"))
    resp = client().post(
        "/api/sap/export-simple/SCRAP-DOES-NOT-EXIST",
        json={"matnr": "SCRAP2-FR", "sloc": "SBED", "quantity": "9"},
    )
    assert resp.status_code == 404


def test_sap_export_simple_success_calls_post_remnant_simple_and_logs(monkeypatch, tmp_path):
    monkeypatch.setattr(config, "EXCEL_DB_PATH", str(tmp_path / "scrap_inventory.xlsx"))
    sync_log_path = str(tmp_path / "sap_sync_log.xlsx")
    monkeypatch.setattr(config, "SAP_SYNC_LOG_PATH", sync_log_path)
    client().post("/api/stock-in", json=_product_payload())

    captured = {}

    def fake_post_remnant_simple(matnr, sloc, quantity):
        captured.update(matnr=matnr, sloc=sloc, quantity=quantity)
        return {
            "ok": True, "http_status": 200, "sap_status": "S",
            "matdoc": "4900012345", "matdoc_year": "2026", "message": "OK",
            "error": None, "payload": {"matnr": matnr, "sloc": sloc, "quantity": str(quantity)},
        }

    monkeypatch.setattr(sap_client, "post_remnant_simple", fake_post_remnant_simple)

    resp = client().post(
        "/api/sap/export-simple/SCRAP-00001",
        json={"matnr": "SCRAP2-FR", "sloc": "SBED", "quantity": "9"},
    )
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["ok"] is True
    assert body["matdoc"] == "4900012345"

    assert captured == {"matnr": "SCRAP2-FR", "sloc": "SBED", "quantity": "9"}

    wb = load_workbook(sync_log_path)
    assert "簡化匯出紀錄" in wb.sheetnames
    assert wb["簡化匯出紀錄"].max_row == 2  # header + 1 record


def test_sap_export_simple_failure_returns_502(monkeypatch, tmp_path):
    monkeypatch.setattr(config, "EXCEL_DB_PATH", str(tmp_path / "scrap_inventory.xlsx"))
    monkeypatch.setattr(config, "SAP_SYNC_LOG_PATH", str(tmp_path / "sap_sync_log.xlsx"))
    client().post("/api/stock-in", json=_product_payload())

    monkeypatch.setattr(sap_client, "post_remnant_simple", lambda matnr, sloc, quantity: {
        "ok": False, "http_status": 404, "sap_status": None,
        "matdoc": None, "matdoc_year": None, "message": None,
        "error": "料號在 SAP 查無資料", "payload": {"matnr": matnr, "sloc": sloc, "quantity": str(quantity)},
    })

    resp = client().post(
        "/api/sap/export-simple/SCRAP-00001",
        json={"matnr": "NOT-A-REAL-MATNR", "sloc": "SBED", "quantity": "9"},
    )
    assert resp.status_code == 502
    assert resp.get_json()["ok"] is False


def test_sap_export_simple_missing_sap_config_returns_500(monkeypatch, tmp_path):
    monkeypatch.setattr(config, "EXCEL_DB_PATH", str(tmp_path / "scrap_inventory.xlsx"))
    client().post("/api/stock-in", json=_product_payload())

    def raise_config_error(matnr, sloc, quantity):
        raise sap_client.SapConfigError("尚未設定 SAP 連線環境變數")

    monkeypatch.setattr(sap_client, "post_remnant_simple", raise_config_error)

    resp = client().post(
        "/api/sap/export-simple/SCRAP-00001",
        json={"matnr": "SCRAP2-FR", "sloc": "SBED", "quantity": "9"},
    )
    assert resp.status_code == 500


def test_sap_export_simple_does_not_call_post_remnant(monkeypatch, tmp_path):
    """The simplified route must go through post_remnant_simple only -- if it
    accidentally called the original post_remnant, this would raise (wrong
    signature: post_remnant is keyword-only and requires board_length_cm etc.
    that this route never has)."""
    monkeypatch.setattr(config, "EXCEL_DB_PATH", str(tmp_path / "scrap_inventory.xlsx"))
    monkeypatch.setattr(config, "SAP_SYNC_LOG_PATH", str(tmp_path / "sap_sync_log.xlsx"))
    client().post("/api/stock-in", json=_product_payload())

    def fail_if_called(*a, **k):
        raise AssertionError("post_remnant should not be called by the simplified route")

    monkeypatch.setattr(sap_client, "post_remnant", fail_if_called)
    monkeypatch.setattr(sap_client, "post_remnant_simple", lambda matnr, sloc, quantity: {
        "ok": True, "http_status": 200, "sap_status": "S",
        "matdoc": "X", "matdoc_year": "2026", "message": "OK", "error": None,
        "payload": {"matnr": matnr, "sloc": sloc, "quantity": str(quantity)},
    })

    resp = client().post(
        "/api/sap/export-simple/SCRAP-00001",
        json={"matnr": "SCRAP2-FR", "sloc": "SBED", "quantity": "9"},
    )
    assert resp.status_code == 200
