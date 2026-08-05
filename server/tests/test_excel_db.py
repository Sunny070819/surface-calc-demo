import json

from openpyxl import load_workbook

import excel_db


def _sample_kwargs(**overrides):
    base = dict(
        batch="BATCH-001",
        source_filename="test.png",
        source_type="image",
        contour_method="自動偵測",
        area_cm2=812.4,
        polygon_cm=[{"x": 0, "y": 0}, {"x": 10, "y": 0}, {"x": 10, "y": 10}],
        length_cm=10.0,
        width_cm=10.0,
        fits_by_product={"A": {"fits": 3, "orientation_used_deg": 0}, "B": {"fits": 2, "orientation_used_deg": 90}},
        material="防火",
        grid_cell_size_cm=1.0,
        verdict="合格入庫",
    )
    base.update(overrides)
    return base


def test_ensure_workbook_creates_header_only_once(tmp_excel_path):
    path1 = excel_db.ensure_workbook(tmp_excel_path)
    path2 = excel_db.ensure_workbook(tmp_excel_path)
    assert path1 == path2 == tmp_excel_path
    wb = load_workbook(tmp_excel_path)
    ws = wb.active
    assert ws.max_row == 1
    assert [c.value for c in ws[1]] == excel_db.HEADERS


def test_append_two_records(tmp_excel_path):
    row1 = excel_db.append_record(path=tmp_excel_path, **_sample_kwargs())
    row2 = excel_db.append_record(path=tmp_excel_path, **_sample_kwargs(material="防水"))

    assert row1["餘料編號"] != row2["餘料編號"]
    assert row1["餘料編號"] == "SCRAP-00001"
    assert row2["餘料編號"] == "SCRAP-00002"

    wb = load_workbook(tmp_excel_path)
    ws = wb.active
    assert ws.max_row == 3  # header + 2 rows
    assert ws.cell(row=2, column=1).value == "SCRAP-00001"
    assert ws.cell(row=3, column=1).value == "SCRAP-00002"
    # 可容納A的數量 / 可容納B的數量 columns
    a_col = excel_db.HEADERS.index("可容納A的數量") + 1
    b_col = excel_db.HEADERS.index("可容納B的數量") + 1
    assert ws.cell(row=2, column=a_col).value == 3
    assert ws.cell(row=2, column=b_col).value == 2


def test_append_record_stores_design_provenance_fields(tmp_excel_path):
    row = excel_db.append_record(
        path=tmp_excel_path,
        **_sample_kwargs(sheet_name="SHEET-1", sheet_area_cm2=10800.0, product_area_cm2=900.0, product_fit_count=9),
    )
    shape_desc = json.loads(row["形狀描述/輪廓座標"])
    assert shape_desc["sheet_name"] == "SHEET-1"
    assert shape_desc["sheet_area_cm2"] == 10800.0
    assert shape_desc["product_area_cm2"] == 900.0
    assert shape_desc["product_fit_count"] == 9


def test_only_qualifying_records_written_by_caller_convention(tmp_excel_path):
    # excel_db itself just appends whatever it's given -- the "only qualifying
    # records get written" rule lives in app.py's stock-in route, which simply
    # never calls append_record for a non-qualifying evaluate() result.
    excel_db.append_record(path=tmp_excel_path, **_sample_kwargs())
    wb = load_workbook(tmp_excel_path)
    assert wb.active.max_row == 2


def test_get_record_returns_none_when_workbook_missing(tmp_excel_path):
    assert excel_db.get_record("SCRAP-00001", path=tmp_excel_path) is None


def test_get_record_finds_matching_row(tmp_excel_path):
    excel_db.append_record(path=tmp_excel_path, **_sample_kwargs())
    row = excel_db.get_record("SCRAP-00001", path=tmp_excel_path)
    assert row is not None
    assert row["餘料編號"] == "SCRAP-00001"
    assert row["材質"] == "防火"


def test_get_record_returns_none_for_unknown_id(tmp_excel_path):
    excel_db.append_record(path=tmp_excel_path, **_sample_kwargs())
    assert excel_db.get_record("SCRAP-99999", path=tmp_excel_path) is None
