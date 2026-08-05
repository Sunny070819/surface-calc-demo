from openpyxl import load_workbook

import sap_sync_log


def _sample_result(**overrides):
    base = dict(
        ok=True, http_status=200, sap_status="S", guid="G-1",
        batch="0000000243", bin_zone="B-1", remnant_area="10800.000",
        message="OK MatDoc:4900011979", error=None,
    )
    base.update(overrides)
    return base


def test_ensure_workbook_creates_header_only_once(tmp_path):
    path = str(tmp_path / "sap_sync_log.xlsx")
    path1 = sap_sync_log.ensure_workbook(path)
    path2 = sap_sync_log.ensure_workbook(path)
    assert path1 == path2 == path
    wb = load_workbook(path)
    ws = wb.active
    assert ws.max_row == 1
    assert [c.value for c in ws[1]] == sap_sync_log.HEADERS


def test_append_sync_record_success(tmp_path):
    path = str(tmp_path / "sap_sync_log.xlsx")
    row = sap_sync_log.append_sync_record(
        scrap_id="SCRAP-00001", source_id="SCRAP-00001", matnr="SCRAP-WP",
        board_length_cm=240.0, board_width_cm=120.0, product_area_cm2=18000.0,
        result=_sample_result(), path=path,
    )
    assert row["結果"] == "成功"
    assert row["SAP回傳guid"] == "G-1"
    assert row["SAP回傳batch"] == "0000000243"
    assert row["SAP回傳bin_zone"] == "B-1"
    assert row["SAP回傳remnant_area"] == "10800.000"

    wb = load_workbook(path)
    assert wb.active.max_row == 2


def test_append_sync_record_failure_records_error_message(tmp_path):
    path = str(tmp_path / "sap_sync_log.xlsx")
    row = sap_sync_log.append_sync_record(
        scrap_id="SCRAP-00001", source_id="SCRAP-00001", matnr="SCRAP-WP",
        board_length_cm=240.0, board_width_cm=120.0, product_area_cm2=99999.0,
        result=_sample_result(
            ok=False, sap_status="E", guid=None, batch=None, bin_zone=None,
            remnant_area=None, message=None, error="成品面積超過原紙板",
        ),
        path=path,
    )
    assert row["結果"] == "失敗"
    assert row["錯誤訊息"] == "成品面積超過原紙板"
    assert row["matnr"] == "SCRAP-WP"


def _sample_simple_result(**overrides):
    base = dict(
        ok=True, http_status=200, sap_status="S",
        matdoc="4900012345", matdoc_year="2026", message="OK", error=None,
    )
    base.update(overrides)
    return base


def test_ensure_simple_workbook_adds_sheet_without_touching_original(tmp_path):
    path = str(tmp_path / "sap_sync_log.xlsx")
    sap_sync_log.ensure_workbook(path)  # simulate the original flow having already created the file
    sap_sync_log.ensure_simple_workbook(path)

    wb = load_workbook(path)
    assert "SAP同步紀錄" in wb.sheetnames
    assert "簡化匯出紀錄" in wb.sheetnames
    assert wb["SAP同步紀錄"].max_row == 1  # untouched: only the header row
    assert [c.value for c in wb["簡化匯出紀錄"][1]] == sap_sync_log.SIMPLE_HEADERS


def test_append_simple_sync_record_success(tmp_path):
    path = str(tmp_path / "sap_sync_log.xlsx")
    row = sap_sync_log.append_simple_sync_record(
        scrap_id="SCRAP-00001", matnr="SCRAP2-FR", sloc="SBED", quantity="9",
        result=_sample_simple_result(), path=path,
    )
    assert row["結果"] == "成功"
    assert row["SAP回傳matdoc"] == "4900012345"
    assert row["SAP回傳matdoc_year"] == "2026"

    wb = load_workbook(path)
    assert wb["簡化匯出紀錄"].max_row == 2
    assert wb.active.title == "SAP同步紀錄"  # active sheet unaffected


def test_append_simple_sync_record_failure_records_error_message(tmp_path):
    path = str(tmp_path / "sap_sync_log.xlsx")
    row = sap_sync_log.append_simple_sync_record(
        scrap_id="SCRAP-00002", matnr="SCRAP2-WP", sloc="CUBE", quantity="3",
        result=_sample_simple_result(
            ok=False, sap_status="E", matdoc=None, matdoc_year=None,
            message=None, error="料號在 SAP 查無資料",
        ),
        path=path,
    )
    assert row["結果"] == "失敗"
    assert row["錯誤訊息"] == "料號在 SAP 查無資料"
    assert row["sloc"] == "CUBE"
